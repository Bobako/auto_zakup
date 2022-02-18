import copy
import datetime
import os
import sys
import time
import traceback
from functools import wraps

from flask import Flask, render_template, request, flash, redirect, url_for
from sqlalchemy import exc, desc, asc
from flask_login import LoginManager, login_user, login_required, current_user, logout_user
import openpyxl

from FlaskApp.db_handler import *
from FlaskApp.bot import Bot
from FlaskApp.cfg import *



app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD
db = Handler()
login_manager = LoginManager()
login_manager.init_app(app)
bot = Bot(db.session)

DAYS = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]


def access_log(f):
    @wraps(f)
    def wrap(*args, **kwargs):
        start = time.perf_counter()
        r = f(*args, **kwargs)
        fin = time.perf_counter()
        with open(LOGS_PATH, "a") as file:
            file.write(f"{datetime.datetime.now().strftime('%d.%m %H:%M')} - {f.__name__}, time: {fin-start:0.4f}\n")
        return r
    return wrap

@login_manager.unauthorized_handler
def unauthorized():
    return redirect(url_for("login") + f"?next={request.url}")


@login_manager.user_loader
def load_user(user_id):
    try:
        user = db.session.query(User).filter(User.id == user_id).one()
        return user
    except exc.NoResultFound:
        return None


@app.route("/login", methods=['post', 'get'])
@access_log
def login():
    if current_user:
        db_logout(current_user)
    if request.method == "POST":
        code = request.form.get("code")
        try:
            user = db.session.query(User).filter(User.code == code).one()
        except exc.NoResultFound:
            flash("Некорректный код")
        else:
            db_login(user)
            next_page = request.args.get('next')
            if not next_page:
                next_page = url_for('index')
            return redirect(next_page)

    return render_template("login.html")



@app.route("/users", methods=['post', 'get'])
@login_required
@access_log
def users_page():
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))
    if request.method == 'POST':
        users = parse_forms(request.form, ["is_admin"])
        users[str(current_user.id)]["is_admin"] = True
        try:
            users[str(current_user.id)].pop("delete")
        except KeyError:
            pass
        pins = []
        correct = True
        for user in users.values():
            if user["code"] in pins:
                flash("Пин коды должны быть уникальны")
                correct = False
                break
            else:
                pins.append(user["code"])

        if correct:
            for user in users.values():
                facilities = []
                for key in user.keys():
                    if 'fid_' in key:
                        print(key)
                        facility_id = int(key.replace('fid_', ''))
                        facilities.append(db.session.query(Facility).filter(Facility.id == facility_id).one())
                user["facilities"] = facilities
                for facility in facilities:
                    user.pop(f"fid_{facility.id}")

            update_objs(users, User)
    return render_template("users.html", users=db.session.query(User).all(),
                           facilities=db.session.query(Facility).all())


@app.route("/facilities", methods=['post', 'get'])
@login_required
@access_log
def facilities_page():
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))
    if request.method == 'POST':
        facilities = parse_forms(request.form)
        names = []
        correct = True
        for facility in facilities.values():
            if facility["name"] in names:
                flash("Названия должны быть уникальным")
                correct = False
                break
            else:
                names.append(facility["name"])
        if correct:
            update_objs(facilities, Facility)
    return render_template("facilities.html", facilities=db.session.query(Facility).all())


@app.route("/units", methods=['post', 'get'])
@login_required
@access_log
def units_page():
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))
    if request.method == "POST":
        units = parse_forms(request.form)

        deses = []
        correct = True
        for unit in units.values():
            if unit["designation"] in deses:
                flash("Обозначения должны быть уникальны")
                correct = False
                break
            else:
                deses.append(unit["designation"])
        if correct:
            update_objs(units, Unit)
    return render_template("units.html", units=db.session.query(Unit).all())


@app.route("/products", methods=['post', 'get'])
@login_required
@access_log
def products_page():
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))
    if request.method == "POST":
        products = parse_forms(request.form, ["alco"])
        if "products" in products:
            products.pop("products")
        strange = []
        for key, product in products.items():
            if "unit_id" not in product:
                strange.append(key)
            else:
                if not int(product["unit_id"]):
                    product["unit_id"] = None
        for strange_ in strange:
            products.pop(strange_)

        update_objs(products, Product)
        file = request.files["products:file"]
        if file:
            filename = "file.xlsx"
            try:
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                els = parse_file(UPLOAD + "/" + filename)
            except Exception as ex:
                msg, type_, tb = sys.exc_info()
                print(f"Error: {msg}, {type_}")
                traceback.print_tb(tb)
                error = ex
                els = None
            else:
                error = False
            finally:
                os.remove(UPLOAD + "/" + filename)
                return products_import(error, els)

    return render_template("products.html", products=db.session.query(Product).order_by(Product.name).all(),
                           units=db.session.query(Unit).all())


def products_import(error, els=None):
    if not els:
        els = []
    already_exist = []
    not_found = []
    if not error:
        for el in els:
            vendor = db.session.query(Vendor).filter(Vendor.name == el["vendor_name"]).first()
            product = db.session.query(Product).filter(Product.name == el["product_name"]).first()
            unit = db.session.query(Unit).filter(Unit.designation == el["unit_designation"]).first()
            facilities = []
            for f_name in el["facilities_names"]:
                facilities.append(db.session.query(Facility).filter(Facility.name == f_name).first())

            if not vendor:
                not_found.append(f"{el['product_name']} -  нет поставщика '{el['vendor_name']}'")
                continue

            if product:
                if product in vendor.products:
                    already_exist.append(product.name)
                    continue

            if not unit:
                not_found.append(f"{el['product_name']} -  нет ед. измерения '{el['unit_designation']}'")
                continue

            if not product:
                product = Product(el["product_name"], unit.id, False)
                db.session.add(product)
            vendor.products.append(product)
            for facility in facilities:
                if facility:
                    if facility not in vendor.facilities:
                        vendor.facilities.append(facility)
            db.session.commit()
    l = len(els) - len(already_exist) - len(not_found)
    return render_template("products_import.html", error=error, already_exist=already_exist, not_found=not_found, l=l)


def parse_file(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    els = []
    row = 0
    blanks = 0
    while True:
        el = {}
        row += 1
        if not sheet.cell(row, 1).value:
            blanks += 1
            if blanks > 3:
                break
            continue
        blanks = 0
        el["vendor_name"] = sheet.cell(row, 1).value.strip() if sheet.cell(row, 1).value else ''
        el["product_name"] = sheet.cell(row, 2).value.strip() if sheet.cell(row, 2).value else ''
        el["unit_designation"] = sheet.cell(row, 3).value.strip() if sheet.cell(row, 3).value else ''
        col = 4
        el["facilities_names"] = []
        while sheet.cell(row, col).value:
            el["facilities_names"].append(sheet.cell(row, col).value.strip() if sheet.cell(row, col).value else '')
            col += 1
        els.append(el)
    wb.close()
    return els


@app.route("/vendors", methods=['post', 'get'])
@login_required
@access_log
def vendors_page():
    start = time.perf_counter()
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))

    if request.method == "POST":
        vendors = parse_forms(request.form)

        for id_, vendor in vendors.items():
            vendor["products"] = []
            for product_id in vendor['products_ids'].split(":"):
                if product_id:
                    if (product := db.session.query(Product).filter(
                            Product.id == product_id).one()) not in vendor["products"]:
                        vendor["products"].append(product)
            schedule = ''
            for key in vendor:
                if "schedule" in key:
                    schedule += key[-1]
            if schedule != '':
                vendor["schedule"] = int(schedule)
            else:
                vendor["schedule"] = None
            vendor["facilities"] = []
            for facility_id in vendor['facilities_ids'].split(":"):
                if facility_id:
                    if (facility := db.session.query(Facility).filter(
                            Facility.id == facility_id).one()) not in vendor["facilities"]:
                        vendor["facilities"].append(facility)
            vendor.pop('products_ids')
            vendor.pop('facilities_ids')
            for i in range(7):
                if "schedule" + str(i) in vendor:
                    vendor.pop("schedule" + str(i))
            try:
                vendor["tg_id"] = int(vendor["tg_id"])
            except ValueError:
                vendor["tg_id"] = ''

            one_vendor = {id_: vendor}
            if 'delete' in vendor:
                update_orders(id_, vendor["products"], vendor["facilities"], True)
                update_objs(one_vendor, Vendor)
            else:
                vendor_id = update_objs(one_vendor, Vendor)[0]
                update_orders(vendor_id, vendor["products"], vendor["facilities"])
    fin = time.perf_counter()
    with open(LOGS_PATH, "a") as file:
        file.write(f"{datetime.datetime.now().strftime('%d.%m %H:%M')} - vendors without rendering, time: {fin - start:0.4f}\n")
    return render_template("vendors.html", vendors=db.session.query(Vendor).all(),
                           facilities=db.session.query(Facility).all(),
                           products=db.session.query(Product).all(), days=DAYS)


def update_orders(vendor_id, products, facilities, delete=False):
    if not delete:
        for facility in facilities:
            for order in facility.orders:
                ordered = order.products
                ordered_products = [ordered_.product for ordered_ in ordered]
                for product in products:
                    if product not in ordered_products:
                        db.session.add(OrderedProduct(product.id, 0, vendor_id, order.id, product.unit_id, True))
    else:
        products = db.session.query(OrderedProduct).filter(OrderedProduct.vendor_id == vendor_id).all()
        for product in products:
            order = product.order
            op = order.products
            op = [prod for prod in op if prod.id != product.id]
            order.products = op
            db.session.delete(product)
            db.session.commit()
    db.session.commit()


@app.route("/preview", methods=['post', 'get'])
@login_required
@access_log
def preview():
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))
    order_id = request.args.get('id')
    order = db.session.query(Order).filter(Order.id == order_id).one()
    if request.method == "POST":
        if not request.form.get("order:confirm"):
            return redirect(url_for("index"))

        msgs = parse_forms(request.form)
        msgs.pop("order")
        order.status = "ORDERED"
        order.sent_date = datetime.datetime.now()
        send_order(order, msgs)
        db.session.commit()
        return redirect(url_for("index"))

    order = previewable_order(order)
    vendors = []
    for product in order.products:
        if product.amount:
            if product.vendor not in vendors:
                vendors.append(product.vendor)

    able = True
    for vendor in vendors:
        if not vendor.tg_id:
            able = False
    return render_template("preview.html", order=order, vendors=vendors, able=able)


def previewable_order(order):
    if order.status == "ADDITIONAL" and order.copy_id:
        old_order = db.session.query(Order).filter(Order.id == order.copy_id).one()
        new_order = Order(order.user_id, order.facility_id, order.date, order.status, False)
        db.session.add(new_order)
        db.session.commit()
        for product, old_product in zip(order.products, old_order.products):
            new_product = OrderedProduct(product.product_id, product.amount - old_product.amount, product.vendor_id,
                                         new_order.id)
            db.session.add(new_product)
        db.session.commit()
        return new_order
    return order


def send_order(order, msgs):  # msgs - {vendor.id:msg}
    copy_order(order)
    for product in order.products:
        if product.amount:
            product.product.orders_count += 1
            product.product.orders_amount += product.amount
    db.session.commit()
    for vendor_id, msg in msgs.items():
        id_ = db.session.query(Vendor).filter(Vendor.id == vendor_id).one().tg_id
        bot.noti_vendor(int(id_), msg['msg'])


def copy_order(order):
    if order.copy_id:
        for product in db.session.query(Order).filter(Order.id == order.copy_id).one().products:
            db.session.delete(product)
    else:
        new_order = Order(None, None, None, None, False)
        db.session.add(new_order)
        db.session.commit()
        order.copy_id = new_order.id
    for product in order.products:
        new_product = OrderedProduct(product.product.id, product.amount, product.vendor_id, order.copy_id,
                                     product.unit_id, product.official)
        db.session.add(new_product)
    db.session.commit()


@app.route("/order_format", methods=['post', 'get'])
@login_required
@access_log
def order_format():
    user = current_user
    if not user.is_admin:
        return redirect(url_for("index"))
    if request.method == "POST":
        if request.form.get("default"):
            msg = DEFAULT_ORDER_FORMAT
        elif request.form.get("confirm"):
            msg = request.form.get("msg")
        msg = msg.replace("}\n", "}")
        db.session.query(MSGFormat).one().msg = msg
        db.session.commit()
        with open(FORMAT_PATH, "w", encoding='utf-8') as file:
            file.write(msg)
    msg = db.session.query(MSGFormat).one().msg
    return render_template("order_format.html", msg=msg)


@app.route("/", methods=['post', 'get'])
@login_required
@access_log
def index():
    user = current_user
    if request.method == 'POST':
        order = parse_forms(request.form, ["official"])
        order_id = list(order["order"].keys())[1]
        if order_id == "new":
            order_id = create_order(order, user)
        else:
            order_id = update_order(order, order_id, user)

        if user.is_admin and order_id:
            return redirect(url_for("preview") + f"?id={order_id}")

    orders = get_orders(user)

    old_orders = get_old_orders(user)

    vendors = db.session.query(Vendor).all()
    products = db.session.query(Product).all()

    col1 = max([len(str(vendor.name)) for vendor in vendors] + [5])
    col2 = max([len(str(product.name)) for product in products] + [5])
    col3 = 3
    col4 = max([len(str(product.unit.designation)) if product.unit else 2 for product in products] + [2])
    facility_id = request.args.get('fid')
    if not facility_id:
        if user.facilities:
            facility_id = user.facilities[0].id
        else:
            facility_id = 0
    return render_template("orders.html", user=user, orders=orders, old_orders=old_orders,
                           products=products, days=DAYS,
                           vendors=vendors, facilities=db.session.query(Facility).all(),
                           col1=col1, col2=col2, col3=col3, col4=col4, facility_id=int(facility_id),
                           deleted_orders=get_deleted_orders(user), units=db.session.query(Unit).all())


def get_orders(user):
    if user.is_admin:
        orders = db.session.query(Order).order_by(desc(Order.create_date)).filter(Order.status != "ORDERED").filter(
            Order.display == True).filter(Order.deleted == False).all()
    else:
        orders = db.session.query(Order).order_by(desc(Order.create_date)).filter(
            Order.facility_id.in_([f.id for f in user.facilities])).filter(Order.status != "ORDERED").filter(
            Order.display == True).filter(Order.deleted == False).all()
    return orders


def get_old_orders(user):
    if user.is_admin:
        orders = db.session.query(Order).order_by(desc(Order.create_date)).filter(Order.status == "ORDERED").filter(
            Order.display == True).filter(Order.deleted == False).all()
    else:
        orders = db.session.query(Order).order_by(desc(Order.create_date)).filter(
            Order.facility_id.in_([f.id for f in user.facilities])).filter(Order.status == "ORDERED").filter(
            Order.display == True).filter(Order.deleted == False).all()
    return orders


def get_deleted_orders(user):
    if user.is_admin:
        orders = db.session.query(Order).order_by(desc(Order.delete_date)).filter(
            Order.display == True).filter(Order.deleted == True).all()
    else:
        orders = db.session.query(Order).order_by(desc(Order.delete_date)).filter(
            Order.facility_id.in_([f.id for f in user.facilities])).filter(
            Order.display == True).filter(Order.deleted == True).all()
    for order in orders:
        if order.delete_date > datetime.datetime.now() + datetime.timedelta(days=30):
            for product in order.products:
                db.session.delete(product)
            cp_order = db.session.query(Order).filter(Order.id == order.copy_id).one()
            for product in cp_order.products:
                db.session.delete(product)
            db.session.delete(cp_order)
            db.session.delete(order)
    db.session.commit()
    return orders


def create_order(order, user):
    facility_id = order["order"]["facility_id"]
    order.pop("order")
    db.session.add(Order(user.id, facility_id))
    db.session.commit()
    order_id = db.session.query(Order).order_by(desc(Order.id)).first().id
    parse_order_products(order, order_id)
    if not user.is_admin:
        bot.noti_admin(
            f"{user.name} {user.surname} {'(' + db.session.query(Facility).filter(Facility.id == facility_id).one().name + ')'} cоздал заказ "
            f"#{order_id}", db.session.query(Facility).filter(Facility.id == facility_id).one().tg_id)
    return order_id


def parse_order_products(order, order_id):
    for product_id, product_dict in order.items():

        try:
            product_dict["amount"] = float(product_dict["amount"])
        except ValueError:
            product_dict["amount"] = 0.0

        if "NEW" in product_id:
            db.session.add(OrderedProduct(product_dict["product_id"], product_dict["amount"], product_dict["vendor_id"],
                                          order_id, product_dict["unit_id"], product_dict["official"]))
        else:
            product = db.session.query(OrderedProduct).filter(OrderedProduct.id == product_id).one()
            product.amount = product_dict["amount"]
            product.unit_id = product_dict["unit_id"]
            product.vendor_id = product_dict["vendor_id"]
            product.official = product_dict["official"]
    db.session.commit()


def update_order(order, order_id, user):
    if "Удалить" in order["order"].values():
        order_obj = db.session.query(Order).filter(Order.id == order_id).one()
        order_obj.deleted = True
        order_obj.delete_date = datetime.datetime.now()
        db.session.commit()
        return None
    if "Восстановить" in order["order"].values():
        order_obj = db.session.query(Order).filter(Order.id == order_id).one()
        order_obj.deleted = False
        order_obj.delete_date = None
        db.session.commit()
        return None
    order.pop("order")
    parse_order_products(order, order_id)
    order_obj = db.session.query(Order).filter(Order.id == order_id).one()
    if order_obj.status == "ORDERED":
        order_obj.status = "ADDITIONAL"
    order_obj.create_date = datetime.datetime.now()
    db.session.commit()
    if not user.is_admin:
        bot.noti_admin(
            f"{user.name} {user.surname} {'(' + db.session.query(Facility).filter(Facility.id == order_obj.facility_id).one().name + ')'} изменил заказ "
            f"#{order_obj.id}", db.session.query(Facility).filter(Facility.id == order_obj.facility_id).one().tg_id)
    return order_id


@app.route("/api/available_products", methods=['get'])
def get_available_products():
    facility_id = request.args.get('id')
    return render_template('order_form.html',
                           vendors=db.session.query(Facility).filter(Facility.id == facility_id).one().vendors,
                           units=db.session.query(Unit).all())


@app.route("/api/formatted_order", methods=['get'])
def formatted_order():
    order_id = request.args.get('oid')
    order = db.session.query(Order).filter(Order.id == order_id).one()
    vendor = db.session.query(Vendor).filter(Vendor.id == request.args.get('vid')).one()
    return render_template('formatted_order.html',
                           order=order, vendor=vendor).replace("\n", "").replace("<br>", "\n")


@app.route("/api/search", methods=['get'])
def search():
    s = request.args.get('s')
    if s:
        products = db.session.query(Product).filter(Product.name.ilike(f"%{s}%")).order_by(Product.name).all()
        res = ""
        for product in products:
            res += f'<button class="hint" type="button" onclick="addEl(this, {product.id})">{product.name}</button><br>'
        return res
    return ''

def merge_sort(stats: list, func):
    l = len(stats)
    if l < 2:
        return stats
    stats1 = merge_sort(stats[:int(l / 2)], func)
    stats2 = merge_sort(stats[int(l / 2):], func)

    i1 = 0
    i2 = 0
    stats = []
    while i1 < len(stats1) or i2 < len(stats2):
        if i1 == len(stats1):
            stats.append(stats2[i2])
            i2 += 1
        elif i2 == len(stats2):
            stats.append(stats1[i1])
            i1 += 1
        else:
            if func(stats1[i1], stats2[i2]) < 0:
                stats.append(stats1[i1])
                i1 += 1
            else:
                stats.append(stats2[i2])
                i2 += 1
    return stats


def less(stats1, stats2):
    return stats1[2] - stats2[2]


def more(stats1, stats2):
    return stats2[2] - stats1[2]


sf = less


@app.route("/stats", methods=['post', 'get'])
@login_required
@access_log
def stats_page():
    global sf
    if request.method == "POST":
        if sf == less:
            sf = more
        else:
            sf = less
    stats = []
    for vendor in db.session.query(Vendor).all():
        for product in vendor.products:
            if orders := db.session.query(OrderedProduct).filter(OrderedProduct.product_id == product.id).all():

                stats.append(
                    [f"{product.name} ({vendor.name})", sum([bool(order.amount) for order in orders]),
                     sum([order.amount for order in orders]),
                     product.unit.designation if product.unit else 'Ед.'])
            else:
                stats.append([f"{product.name} ({vendor.name})", 0, 0, product.unit.designation if product.unit else 'Ед.'])
    return render_template("stats.html",
                           products=merge_sort(stats, sf))


@app.route("/logs", methods=['post', 'get'])
@login_required
@access_log
def logs_page():
    return


@app.route("/notifications", methods=['post', 'get'])
@login_required
@access_log
def noti_page():
    if request.method == "POST":
        noti = db.session.query(Noti).one()
        noti.tg_id = request.form.get("tg_id")
        noti.send = bool(request.form.get("send"))
        db.session.commit()
    return render_template("noti.html", noti=db.session.query(Noti).one())


def db_login(user):
    login_user(user)
    user.is_authenticated = True
    db.session.commit()


def db_logout(user):
    logout_user()
    db.session.commit()


def parse_forms(form, checkboxes=()):
    result = {}

    for key, val in form.lists():
        id_, arg = key.split(":")
        if id_ == "NEW":
            for i, value in enumerate(val):
                id_ = "NEW" + str(i)
                if id_ not in result:
                    result[id_] = dict()
                result[id_][arg] = value
        else:
            val = val[0]
            if id_ not in result:
                result[id_] = dict()

            if val == "on":
                val = True
            result[id_][arg] = val

    for res_dict in result.values():
        for checkbox in checkboxes:
            if checkbox not in res_dict:
                res_dict[checkbox] = False
            else:
                res_dict[checkbox] = True

    return result


def update_objs(dicts, class_, not_nullable="name"):
    ids = []
    for id_, dict_ in dicts.items():
        if "NEW" not in id_:
            obj = db.session.query(class_).filter(class_.id == id_).one()
            if "delete" in dict_:
                db.session.delete(obj)
                continue
            for arg_name, value in dict_.items():
                setattr(obj, arg_name, value)
            ids.append(obj)

        else:
            if not dict_[not_nullable]:
                continue

            obj = class_(**dict_)
            db.session.add(obj)
            ids.append(obj)
    db.session.commit()
    return [obj.id for obj in ids]


if __name__ == '__main__':
    app.run(debug=False)
